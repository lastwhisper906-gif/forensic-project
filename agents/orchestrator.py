"""Forensic Pipeline 오케스트레이터 — 6-Agent 병렬 + Opus 총괄.

Session 4 업데이트:
  - Agent 1-3: Python 사전 계산 메트릭 → LLM 해석 분리
  - Peer 데이터 한 번만 fetch하고 Agent 1-3 공유
  - peer_set.py + quant_metrics.py 연동

Session 5 업데이트:
  - Agent 4: diff_analyzer.py Language Evolution Memo 사전 생성 추가
  - Phase 0.5에서 Agent 1-3 정량 메트릭 + Agent 4 Language Diff Memo 동시 계산

흐름:
  Phase 0: Peer set 조회 + Peer XBRL 데이터 병렬 fetch (Agent 1-3 공유)
  Phase 0.5: Agent 1-3용 정량 메트릭 + Agent 4 Language Diff Memo 사전 계산 (Python)
  Phase 1: 6개 포렌식 에이전트 asyncio.gather 병렬 실행
           Agent 1-3: precomputed 정량 메트릭 context 포함 user prompt
           Agent 4: Language Evolution Memo 포함 user prompt
           Agent 5-6: 기존 방식 (직접 도구 호출)
  Phase 2: Orchestrator (claude-opus-4-6) — 6개 결과 종합,
           Forensic Score 산출, Tier 분류, Next Action 생성

Forensic Score: 0(최악/가장 의심) ~ 100(깨끗한 회계)
  낮을수록 Short 후보에 가까움.

Tier:
  1 = Active Short  (0~30)
  2 = Monitor       (31~55)
  3 = Avoid         (56~70)
  4 = Archive       (71~100)
"""

from __future__ import annotations

import asyncio
import json
import os
import re
import time
from dataclasses import dataclass, field, asdict
from typing import Any

import anthropic

from claude_agent_sdk import (
    AssistantMessage,
    ClaudeAgentOptions,
    ClaudeSDKClient,
    ResultMessage,
    TextBlock,
)

from agents import (
    AGENT_REGISTRY, AGENT_LABELS, FORENSIC_DATA_SERVER,
    build_options, QUANTITATIVE_AGENTS,
)
from peer_set import get_peer_set, get_sector_group
from quant_metrics import agent1_precomputed, agent2_precomputed, agent3_precomputed
from call_metrics import agent5_precomputed
from catalyst_monitor import agent6_precomputed
import data_sources as ds

# diff_analyzer는 선택적 import (anthropic SDK 필요)
try:
    from diff_analyzer import analyze_diff_with_llm
    _DIFF_ANALYZER_AVAILABLE = True
except ImportError:
    _DIFF_ANALYZER_AVAILABLE = False


# ---------------------------------------------------------------------------
# 결과 데이터 구조
# ---------------------------------------------------------------------------

@dataclass
class AgentReport:
    agent: str
    text: str
    summary: dict[str, Any] = field(default_factory=dict)
    elapsed_sec: float = 0.0
    cost_usd: float | None = None
    error: str | None = None


@dataclass
class ForensicResult:
    ticker: str
    elapsed_sec: float
    reports: dict[str, AgentReport]

    def to_dict(self) -> dict[str, Any]:
        return {
            "ticker": self.ticker,
            "elapsed_sec": round(self.elapsed_sec, 2),
            "reports": {k: asdict(v) for k, v in self.reports.items()},
        }

    def to_markdown(self) -> str:
        order = [
            "accruals", "revenue", "capex",
            "tenk_diff", "call_nlp", "catalyst",
        ]
        lines = [
            f"# Forensic Analysis Report: {self.ticker}",
            f"_총 소요시간: {self.elapsed_sec:.2f}s · 6-agent 병렬 + Opus 총괄_",
            f"_분석일: {time.strftime('%Y-%m-%d')}_",
            "",
        ]

        for k in order:
            r = self.reports.get(k)
            if not r:
                continue
            label = AGENT_LABELS.get(k, k)
            lines.append(f"## {label}")
            lines.append(f"_소요 {r.elapsed_sec:.2f}s_")
            lines.append("")
            if r.error:
                lines.append(f"> ERROR: {r.error}")
            else:
                lines.append(r.text.strip())
            lines.append("")

        orch = self.reports.get("Orchestrator")
        if orch:
            lines.append("---")
            lines.append("## 🔬 Orchestrator — Forensic 종합 판단 (Claude Opus)")
            lines.append(f"_소요 {orch.elapsed_sec:.2f}s_")
            lines.append("")
            if orch.error:
                lines.append(f"> ERROR: {orch.error}")
            else:
                lines.append(orch.text.strip())
            lines.append("")

        return "\n".join(lines)


# ---------------------------------------------------------------------------
# SUMMARY_JSON 파싱
# ---------------------------------------------------------------------------

_SUMMARY_RE = re.compile(r"##\s*SUMMARY_JSON\s*\n+\s*(\{.*?\})\s*$", re.DOTALL)


def _extract_summary_json(text: str) -> dict[str, Any]:
    m = _SUMMARY_RE.search(text)
    if not m:
        return {}
    try:
        return json.loads(m.group(1))
    except json.JSONDecodeError:
        return {}


# ---------------------------------------------------------------------------
# Phase 0: Peer Set 조회 + XBRL 데이터 병렬 fetch
# ---------------------------------------------------------------------------

async def fetch_single_xbrl(ticker: str, years: int = 5) -> dict[str, Any] | None:
    """단일 종목 XBRL 데이터 fetch (에러 시 None 반환).

    5년 데이터로 확장 — Sloan 8분기 시계열 계산에 충분한 데이터 확보.
    """
    try:
        return await ds.sec_xbrl_financials(ticker, years=years)
    except Exception as e:
        print(f"    ⚠ XBRL fetch 실패 [{ticker}]: {type(e).__name__}: {e}", flush=True)
        return None


async def fetch_peer_xbrl_batch(
    ticker: str,
    peers: list[str],
    years: int = 3,
) -> dict[str, dict[str, Any]]:
    """Peer 종목 XBRL 데이터 병렬 fetch.

    Args:
        ticker: 기준 종목 (로그용)
        peers:  peer ticker 리스트
        years:  연간 기간 수

    Returns:
        {peer_ticker: xbrl_ts_dict} — fetch 실패한 peer는 포함 안 됨
    """
    if not peers:
        return {}

    print(
        f"    📊 Peer XBRL fetch: {', '.join(peers)} (연간 {years}년) …",
        flush=True,
    )
    tasks = {p: fetch_single_xbrl(p, years=years) for p in peers}
    results = await asyncio.gather(*tasks.values(), return_exceptions=True)

    peer_data: dict[str, dict] = {}
    for peer, result in zip(tasks.keys(), results):
        if isinstance(result, dict):
            peer_data[peer] = result
        # Exception / None → 제외

    print(
        f"    ✓ Peer 데이터 수집: {len(peer_data)}/{len(peers)}개 성공",
        flush=True,
    )
    return peer_data


# ---------------------------------------------------------------------------
# Phase 0.5: Agent 1-3 정량 메트릭 사전 계산
# ---------------------------------------------------------------------------

def compute_precomputed_context(
    agent_key: str,
    ts: dict[str, Any],
    peer_ts_map: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    """에이전트별 사전 계산 컨텍스트 생성.

    Args:
        agent_key:   "accruals" | "revenue" | "capex"
                     ("tenk_diff" 는 compute_diff_memo() 사용)
        ts:          기준 종목 XBRL 시계열
        peer_ts_map: peer 종목 XBRL 시계열 맵

    Returns:
        사전 계산된 메트릭 dict, 또는 None (미지원 에이전트)
    """
    if agent_key == "accruals":
        return agent1_precomputed(ts, peer_ts_map)
    elif agent_key == "revenue":
        return agent2_precomputed(ts, peer_ts_map)
    elif agent_key == "capex":
        return agent3_precomputed(ts, peer_ts_map)
    return None


async def compute_diff_memo(
    ticker: str,
    peer_context: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    """Agent 4 전용: diff_analyzer.py로 Language Evolution Memo 생성.

    SEC EDGAR에서 10-K 섹션을 fetch하고 forensic_engine + diff_analyzer로 분석.

    Args:
        ticker:       종목 코드
        peer_context: 동종업체 비교 데이터 (선택)

    Returns:
        analyze_diff_with_llm() 결과 dict, 또는 None (실패 시)
    """
    if not _DIFF_ANALYZER_AVAILABLE:
        print("    ⚠ diff_analyzer 사용 불가 (anthropic 패키지 필요)", flush=True)
        return None

    if not os.getenv("ANTHROPIC_API_KEY"):
        print("    ⚠ ANTHROPIC_API_KEY 없음 — diff memo 스킵", flush=True)
        return None

    try:
        print(f"    📄 [{ticker}] 10-K 섹션 fetch (current + prior year) …", flush=True)
        # 현재 연도와 전년도 10-K 섹션 병렬 fetch
        sections_to_fetch = [
            "risk_factors",
            "mda",
            "critical_accounting",
            "related_party",
        ]
        current_raw, prior_raw = await asyncio.gather(
            ds.sec_10k_sections(
                ticker,
                prior_year=False,
                sections=sections_to_fetch,
                max_chars_per_section=5000,
            ),
            ds.sec_10k_sections(
                ticker,
                prior_year=True,
                sections=sections_to_fetch,
                max_chars_per_section=5000,
            ),
        )

        # 섹션 이름 정규화 (key mapping)
        _SEC_SECTION_MAP = {
            "risk_factors":          "item_1a_risk_factors",
            "mda":                   "md_and_a",
            "critical_accounting":   "critical_accounting_estimates",
            "related_party":         "related_party_transactions",
        }

        def _extract_sections(raw: dict) -> dict[str, str]:
            secs = raw.get("sections") if isinstance(raw, dict) else {}
            if not isinstance(secs, dict):
                return {}
            out = {}
            for k, v in secs.items():
                normalized = _SEC_SECTION_MAP.get(k, k)
                text = v if isinstance(v, str) else str(v)
                if text and text.strip():
                    out[normalized] = text
            return out

        sections_current = _extract_sections(current_raw)
        sections_prior   = _extract_sections(prior_raw)

        if not sections_current or not sections_prior:
            print(f"    ⚠ [{ticker}] 10-K 섹션 fetch 결과 비어있음", flush=True)
            return None

        # fy 정보 추출
        fy_current = current_raw.get("fy", "FY_CURRENT") if isinstance(current_raw, dict) else "FY_CURRENT"
        fy_prior   = prior_raw.get("fy", "FY_PRIOR")   if isinstance(prior_raw,   dict) else "FY_PRIOR"

        print(
            f"    📝 [{ticker}] Language Diff 분석: "
            f"{len(sections_current)} 섹션 (current) / "
            f"{len(sections_prior)} 섹션 (prior) …",
            flush=True,
        )

        # forensic_engine + diff_analyzer 실행
        from diff_analyzer import render_memo_from_sections
        memo = await render_memo_from_sections(
            ticker=ticker,
            sections_current=sections_current,
            sections_prior=sections_prior,
            fy_current=str(fy_current),
            fy_prior=str(fy_prior),
            peer_context=peer_context or {},
            use_opus_executive=False,
        )

        high = memo.get("high_count", 0)
        cost = memo.get("total_cost_usd", 0)
        print(
            f"    ✓ [{ticker}] Language Diff Memo: HIGH={high}  cost=${cost:.5f}",
            flush=True,
        )
        return memo

    except Exception as e:
        print(f"    ⚠ [{ticker}] diff memo 생성 실패: {type(e).__name__}: {e}", flush=True)
        return None


def build_precomputed_prompt(
    ticker: str,
    precomputed: dict[str, Any],
    peers: list[str],
) -> str:
    """사전 계산 메트릭을 user prompt로 변환 (Agent 1-3용)."""
    peers_str = ", ".join(peers) if peers else "없음"
    metrics_json = json.dumps(precomputed, ensure_ascii=False, indent=2)
    return (
        f"Ticker: {ticker} (US market)\n\n"
        f"## Peer Set (z-score 계산 기준)\n{peers_str}\n\n"
        f"## 사전 계산된 정량 지표 (Python quant_metrics.py 산출)\n"
        f"```json\n{metrics_json}\n```\n\n"
        f"위 지표를 분석의 출발점으로 사용하세요. "
        f"flag=True 항목에 대해 10-K 텍스트 증거를 반드시 확인하세요. "
        f"분석 결과를 Korean markdown + ## SUMMARY_JSON 형식으로 출력하세요."
    )


def build_diff_memo_prompt(
    ticker: str,
    memo: dict[str, Any],
) -> str:
    """Language Evolution Memo를 Agent 4 user prompt로 변환.

    Args:
        ticker: 종목 코드
        memo:   diff_analyzer.analyze_diff_with_llm() 결과 dict

    Returns:
        Agent 4 user_prompt 문자열
    """
    fy_current = memo.get("fy_current", "FY_CURRENT")
    fy_prior   = memo.get("fy_prior",   "FY_PRIOR")
    high_count = memo.get("high_count", 0)
    cost       = memo.get("total_cost_usd", 0)
    findings   = memo.get("findings", [])
    cls_list   = memo.get("classifications", [])

    # findings JSON (간결 버전)
    findings_brief = []
    for f in findings:
        findings_brief.append({
            "no":             f.get("finding_no"),
            "section":        f.get("section_name"),
            "title":          f.get("finding_title"),
            "priority":       f.get("priority"),
            "verdict":        f.get("verdict"),
            "prior_quote":    f.get("fy_prior_quote", "")[:150],
            "current_quote":  f.get("fy_current_quote", "")[:150],
            "impact":         f.get("impact_estimate"),
            "kill_criteria":  f.get("kill_criteria"),
        })

    # Stage 1 분류 요약
    cls_brief = [
        {
            "section": c.get("section_name"),
            "type":    c.get("change_type"),
            "severity": c.get("severity"),
            "summary": c.get("one_liner"),
        }
        for c in cls_list
    ]

    findings_json = json.dumps(findings_brief, ensure_ascii=False, indent=2)
    cls_json      = json.dumps(cls_brief, ensure_ascii=False, indent=2)
    memo_md       = memo.get("memo_markdown", "")[:4000]  # Markdown 일부 포함

    return (
        f"Ticker: {ticker} (US market)\n"
        f"10-K Language Diff: {fy_prior} → {fy_current}\n\n"
        f"## 사전 생성된 Language Evolution Memo\n"
        f"- 분석 모델: Stage1={memo.get('model_used_stage1', 'Haiku')} / "
        f"Stage2={memo.get('model_used_stage2', 'Sonnet')}\n"
        f"- 비용: ${cost:.5f}  |  HIGH findings: {high_count}\n\n"
        f"### Stage 1 섹션 분류 결과\n"
        f"```json\n{cls_json}\n```\n\n"
        f"### Stage 2 Deep Analysis Findings\n"
        f"```json\n{findings_json}\n```\n\n"
        f"### Memo Markdown Preview\n"
        f"{memo_md}\n\n"
        f"---\n"
        f"위 Memo를 검토하고, HIGH priority findings를 sec_10k_sections로 직접 확인하세요. "
        f"최종 종합 판단과 ## SUMMARY_JSON을 Korean markdown 형식으로 출력하세요."
    )


# ---------------------------------------------------------------------------
# 단일 에이전트 실행
# ---------------------------------------------------------------------------

def build_call_metrics_prompt(ticker: str, call_precomputed: dict[str, Any]) -> str:
    """Agent 5 전용: call_metrics 사전 계산 결과 → user prompt."""
    quarters = call_precomputed.get("quarters_analyzed", 0)
    data_src  = call_precomputed.get("data_source", "unknown")
    summary   = call_precomputed.get("summary_text", "데이터 없음")

    # 핵심 지표만 compact JSON으로
    compact = {
        "quarters_analyzed":    quarters,
        "data_source":          data_src,
        "kpi_trends":           call_precomputed.get("kpi_trends", {}),
        "hedging_trend":        call_precomputed.get("hedging_trend", {}),
        "confidence_trend":     call_precomputed.get("confidence_trend", {}),
        "non_gaap_trend":       call_precomputed.get("non_gaap_trend", {}),
        "guidance_quality_latest": call_precomputed.get("guidance_quality_latest", "ABSENT"),
        "flags":                call_precomputed.get("flags", {}),
        "qa_evasions":          call_precomputed.get("qa_evasions", [])[:5],  # 최대 5건
    }
    metrics_json = json.dumps(compact, ensure_ascii=False, indent=2)

    return (
        f"Ticker: {ticker} (US market)\n\n"
        f"## 사전 계산된 Earnings Call 분석 (Python call_metrics.py)\n"
        f"```json\n{metrics_json}\n```\n\n"
        f"## 상세 요약\n{summary}\n\n"
        f"위 사전 계산 결과를 출발점으로 분석하세요. "
        f"flag=True 항목 중 포렌식적으로 중요한 것을 10-K 증거와 교차 확인하세요. "
        f"결과를 Korean markdown + ## SUMMARY_JSON 형식으로 출력하세요."
    )


def build_catalyst_prompt(ticker: str, catalyst_precomputed: dict[str, Any]) -> str:
    """Agent 6 전용: catalyst_monitor 사전 계산 결과 → user prompt."""
    summary = catalyst_precomputed.get("summary_text", "데이터 없음")

    compact = {
        "has_active_catalyst":  catalyst_precomputed.get("has_active_catalyst", False),
        "max_severity":         catalyst_precomputed.get("max_severity", 0),
        "catalyst_probability": catalyst_precomputed.get("catalyst_probability", "NONE"),
        "active_catalysts":     catalyst_precomputed.get("active_catalysts", [])[:6],
        "insider_pattern":      catalyst_precomputed.get("insider_pattern", {}),
        "flags":                catalyst_precomputed.get("flags", {}),
    }
    metrics_json = json.dumps(compact, ensure_ascii=False, indent=2)

    return (
        f"Ticker: {ticker} (US market)\n\n"
        f"## 사전 계산된 Catalyst 이벤트 분류 (Python catalyst_monitor.py)\n"
        f"```json\n{metrics_json}\n```\n\n"
        f"## 상세 요약\n{summary}\n\n"
        f"severity ≥ 80인 이벤트는 sec_8k_items / sec_corresp 호출로 원문을 확인하세요. "
        f"compound_signal=True이면 반드시 상세 보고하세요. "
        f"결과를 Korean markdown + ## SUMMARY_JSON 형식으로 출력하세요."
    )


async def run_single_agent(
    agent_key: str,
    ticker: str,
    precomputed: dict[str, Any] | None = None,
    diff_memo: dict[str, Any] | None = None,
    call_precomputed: dict[str, Any] | None = None,
    catalyst_precomputed: dict[str, Any] | None = None,
) -> AgentReport:
    """ClaudeSDKClient로 포렌식 에이전트 하나 실행.

    Args:
        agent_key:            에이전트 키 (AGENT_REGISTRY)
        ticker:               분석 대상 티커
        precomputed:          Agent 1-3용 사전 계산 메트릭 (None이면 기존 방식)
        diff_memo:            Agent 4용 Language Evolution Memo
        call_precomputed:     Agent 5용 Earnings Call 사전 계산
        catalyst_precomputed: Agent 6용 Catalyst 사전 계산
    """
    started = time.perf_counter()

    if agent_key == "tenk_diff" and diff_memo is not None:
        # Agent 4: Language Evolution Memo 포함 (Session 5 방식)
        user_prompt = build_diff_memo_prompt(ticker, diff_memo)
    elif agent_key == "call_nlp" and call_precomputed is not None:
        # Agent 5: Earnings Call 사전 계산 포함 (Session 6 방식)
        user_prompt = build_call_metrics_prompt(ticker, call_precomputed)
    elif agent_key == "catalyst" and catalyst_precomputed is not None:
        # Agent 6: Catalyst 사전 계산 포함 (Session 6 방식)
        user_prompt = build_catalyst_prompt(ticker, catalyst_precomputed)
    elif precomputed is not None:
        # Agent 1-3: 사전 계산 컨텍스트 포함 (Session 4 방식)
        peers = precomputed.get("peer_tickers", [])
        user_prompt = build_precomputed_prompt(ticker, precomputed, peers)
    else:
        # fallback: LLM이 도구 직접 호출
        user_prompt = (
            f"Ticker: {ticker} (US market)\n\n"
            f"Perform your specialized forensic analysis on this company. "
            f"Call the necessary tools, analyze the data rigorously, "
            f"and output your findings in Korean markdown followed by "
            f"the required `## SUMMARY_JSON` block."
        )

    try:
        options = build_options(agent_key)
        async with ClaudeSDKClient(options=options) as client:
            await client.query(user_prompt)
            chunks: list[str] = []
            cost: float | None = None
            async for message in client.receive_response():
                if isinstance(message, AssistantMessage):
                    for block in message.content:
                        if isinstance(block, TextBlock):
                            chunks.append(block.text)
                elif isinstance(message, ResultMessage):
                    cost = message.total_cost_usd
        full_text = "".join(chunks)
        return AgentReport(
            agent=agent_key,
            text=full_text,
            summary=_extract_summary_json(full_text),
            elapsed_sec=time.perf_counter() - started,
            cost_usd=cost,
        )
    except Exception as e:
        return AgentReport(
            agent=agent_key,
            text="",
            elapsed_sec=time.perf_counter() - started,
            error=f"{type(e).__name__}: {e}",
        )


# ---------------------------------------------------------------------------
# Forensic Score 계산 (sub-score 가중 합산)
# ---------------------------------------------------------------------------

DEFAULT_WEIGHTS: dict[str, float] = {
    "accruals_score":        0.20,
    "revenue_quality_score": 0.20,
    "capex_score":           0.15,
    "tenk_diff_score":       0.20,
    "call_nlp_score":        0.10,
    "catalyst_score":        0.15,
}

SCORE_KEY_MAP: dict[str, str] = {
    "accruals":  "accruals_score",
    "revenue":   "revenue_quality_score",
    "capex":     "capex_score",
    "tenk_diff": "tenk_diff_score",
    "call_nlp":  "call_nlp_score",
    "catalyst":  "catalyst_score",
}


def calculate_forensic_score(
    result: "ForensicResult",
    weights: dict[str, float] | None = None,
) -> dict[str, Any]:
    """6개 sub-score → 가중 합산 Forensic Score + Tier 분류."""
    w = weights or DEFAULT_WEIGHTS
    sub_scores: dict[str, dict] = {}
    weighted_total = 0.0
    weight_used = 0.0

    for agent_key, score_key in SCORE_KEY_MAP.items():
        r = result.reports.get(agent_key)
        weight = w.get(score_key, 0.0)
        if r and not r.error and r.summary:
            score_val = r.summary.get(score_key)
            if score_val is not None:
                score_int = int(score_val)
                weighted_total += score_int * weight
                weight_used += weight
                sub_scores[score_key] = {
                    "score": score_int,
                    "weight": weight,
                    "weighted": round(score_int * weight, 2),
                    "red_flags": r.summary.get("red_flags", []),
                }
            else:
                sub_scores[score_key] = {
                    "score": None,
                    "weight": weight,
                    "weighted": None,
                    "red_flags": r.summary.get("red_flags", []),
                }
        else:
            err = r.error if r else "결과 없음"
            sub_scores[score_key] = {
                "score": None,
                "weight": weight,
                "weighted": None,
                "error": err,
            }

    forensic_score: int | None = None
    if weight_used > 0:
        forensic_score = round(weighted_total / weight_used)

    # has_active_catalyst: Agent 6 SUMMARY_JSON에서 추출
    catalyst_report = result.reports.get("catalyst")
    has_active_catalyst = False
    if catalyst_report and not catalyst_report.error and catalyst_report.summary:
        has_active_catalyst = bool(catalyst_report.summary.get("has_active_catalyst", False))

    tier, tier_label = classify_tier(forensic_score, has_active_catalyst)

    all_red_flags: list[dict] = []
    for agent_key, score_key in SCORE_KEY_MAP.items():
        flags = sub_scores.get(score_key, {}).get("red_flags", [])
        for flag in flags:
            all_red_flags.append({
                "agent": AGENT_LABELS.get(agent_key, agent_key),
                "flag": flag,
                "score": sub_scores[score_key].get("score"),
            })

    return {
        "ticker": result.ticker,
        "forensic_score": forensic_score,
        "tier": tier,
        "tier_label": tier_label,
        "has_active_catalyst": has_active_catalyst,
        "sub_scores": sub_scores,
        "red_flags": all_red_flags,
        "weight_coverage": round(weight_used, 2),
    }


def classify_tier(
    forensic_score: int | None,
    has_active_catalyst: bool = False,
) -> tuple[int, str]:
    """Forensic Score + Active Catalyst → Tier 분류.

    Session 6 업데이트:
      has_active_catalyst=True이면 score 기준 한 단계 상향 (더 위험하게 분류).

    Tier:
      1 = Active Short  (score ≤ 30, 또는 ≤ 55 + catalyst)
      2 = Monitor       (score ≤ 55, 또는 ≤ 70 + catalyst)
      3 = Long Avoid    (score ≤ 70)
      4 = Archive       (score > 70)
    """
    if forensic_score is None:
        return 0, "Unknown"

    # catalyst 보정: score threshold를 25 완화 (더 쉽게 위험 등급)
    effective = forensic_score - (15 if has_active_catalyst else 0)

    if effective <= 30:
        return 1, "Active Short"
    if effective <= 55:
        return 2, "Monitor"
    if effective <= 70:
        return 3, "Long Avoid"
    return 4, "Archive"


# ---------------------------------------------------------------------------
# ForensicResult → DB 호환 dict 변환
# ---------------------------------------------------------------------------

def forensic_result_to_dict(
    result: "ForensicResult",
    skip_orchestrator: bool = False,
    duration_sec: float | None = None,
) -> dict[str, Any]:
    """ForensicResult → db.save_result() 호환 dict."""
    sc   = calculate_forensic_score(result)
    orch = result.reports.get("Orchestrator")
    orch_summary = orch.summary if orch and not orch.error else {}

    agent_reports: list[dict] = []
    for agent_key, score_key in SCORE_KEY_MAP.items():
        r = result.reports.get(agent_key)
        if r:
            agent_reports.append({
                "agent_key": agent_key,
                "score":     r.summary.get(score_key) if r.summary else None,
                "flags":     r.summary.get("red_flags", []) if r.summary else [],
                "summary":   r.text[:500] if r.text else None,
            })

    return {
        "ticker":           result.ticker,
        "forensic_score":   orch_summary.get("forensic_score") or sc["forensic_score"],
        "tier":             orch_summary.get("tier") or sc["tier"],
        "tier_label":       orch_summary.get("tier_label") or sc["tier_label"],
        "short_thesis":     (orch.text[:2000] if orch and not orch.error else None),
        "skip_orchestrator": skip_orchestrator,
        "weights":          DEFAULT_WEIGHTS,
        "error":            None,
        "duration_sec":     duration_sec if duration_sec is not None else result.elapsed_sec,
        "agent_reports":    agent_reports,
    }


# ---------------------------------------------------------------------------
# Orchestrator — Claude Opus 총괄
# ---------------------------------------------------------------------------

async def run_orchestrator(result: "ForensicResult") -> AgentReport:
    """Claude Opus(claude-opus-4-6)로 6개 에이전트 결과를 종합."""
    started = time.perf_counter()

    pre_score = calculate_forensic_score(result)

    sections: list[str] = []
    for agent_key in ["accruals", "revenue", "capex", "tenk_diff", "call_nlp", "catalyst"]:
        r = result.reports.get(agent_key)
        label = AGENT_LABELS.get(agent_key, agent_key)
        if r and not r.error:
            summary_str = json.dumps(r.summary, ensure_ascii=False) if r.summary else "없음"
            body = r.text[:2500] if len(r.text) > 2500 else r.text
            sections.append(f"### [{label}]\n{body}\n\nSUMMARY_JSON: {summary_str}")
        else:
            err = r.error if r else "결과 없음"
            sections.append(f"### [{label}]\n⚠️ 데이터 없음: {err}")

    combined = "\n\n---\n\n".join(sections)

    prompt = f"""당신은 Chanos-Schilit 전통의 포렌식 회계 수석 애널리스트입니다.
6개 전문 에이전트가 **{result.ticker}** (미국 AI 인프라 섹터)를 분석한 결과입니다.
Agent 1-3은 Python 사전 계산 메트릭(Peer z-score 포함)을 기반으로 분석했습니다.

{combined}

---

## 사전 계산된 Forensic Score (참고용)
{json.dumps(pre_score, ensure_ascii=False, indent=2)}

---

다음 단계로 최종 포렌식 판단을 작성하세요.

## 📊 Forensic 종합 분석

### 1. 핵심 회계 품질 이슈 (상위 3~5개)
각 이슈마다: 에이전트 출처 / 수치 근거 / 심각도(HIGH/MEDIUM/LOW) 명시.
Peer z-score가 있는 경우 peer 대비 위치도 언급.

### 2. Tier 분류 근거
사전 계산된 Score와 Tier를 검토하고, 필요시 가중치를 조정하여 최종 Tier를 확정하세요.
조정한 경우 반드시 이유를 설명하세요.

### 3. Short Thesis (있는 경우)
Tier 1 또는 2라면: 어떤 메커니즘으로 주가 하락이 발생할지 1~2개 시나리오.
Tier 3~4라면: 모니터링 중단 근거.

### 4. Next Action
구체적인 날짜 또는 이벤트 기반 액션 아이템 2~3개.

## SUMMARY_JSON
{{
  "ticker": "{result.ticker}",
  "forensic_score": <0-100 정수, 낮을수록 위험>,
  "tier": <1~4 정수>,
  "tier_label": "<Active Short|Monitor|Avoid|Archive>",
  "sub_scores": {{
    "accruals_score":        {{"score": <int|null>, "weight": <float>, "weighted": <float|null>}},
    "revenue_quality_score": {{"score": <int|null>, "weight": <float>, "weighted": <float|null>}},
    "capex_score":           {{"score": <int|null>, "weight": <float>, "weighted": <float|null>}},
    "tenk_diff_score":       {{"score": <int|null>, "weight": <float>, "weighted": <float|null>}},
    "call_nlp_score":        {{"score": <int|null>, "weight": <float>, "weighted": <float|null>}},
    "catalyst_score":        {{"score": <int|null>, "weight": <float>, "weighted": <float|null>}}
  }},
  "red_flags": [
    {{"agent": "<label>", "severity": "HIGH|MEDIUM|LOW", "flag": "<finding>", "evidence": "<数字 or quote>"}}
  ],
  "next_action": {{
    "priority": "HIGH|MEDIUM|LOW",
    "actions": ["<action 1 with date/trigger>", "<action 2>"]
  }},
  "confidence": "HIGH|MEDIUM|LOW"
}}"""

    try:
        client = anthropic.AsyncAnthropic(
            api_key=os.getenv("ANTHROPIC_API_KEY", "")
        )
        message = await client.messages.create(
            model="claude-opus-4-6",
            max_tokens=3500,
            messages=[{"role": "user", "content": prompt}],
        )
        full_text = message.content[0].text
        return AgentReport(
            agent="Orchestrator",
            text=full_text,
            summary=_extract_summary_json(full_text),
            elapsed_sec=time.perf_counter() - started,
        )
    except Exception as e:
        return AgentReport(
            agent="Orchestrator",
            text="",
            elapsed_sec=time.perf_counter() - started,
            error=f"{type(e).__name__}: {e}",
        )


# ---------------------------------------------------------------------------
# Excel 저장
# ---------------------------------------------------------------------------

def export_to_excel(
    results: list["ForensicResult"],
    output_path: str | None = None,
) -> str:
    """포렌식 분석 결과를 reports/forensic_candidates.xlsx에 저장 (누적)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from pathlib import Path
        from datetime import datetime as _dt
    except ImportError:
        raise ImportError("openpyxl 필요: pip install openpyxl")

    if output_path is None:
        base = Path(__file__).parent.parent / "reports"
        base.mkdir(parents=True, exist_ok=True)
        output_path = str(base / "forensic_candidates.xlsx")

    path = Path(output_path)

    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Forensic Candidates"

        headers = [
            "분석일시", "Ticker",
            "Forensic Score", "Tier", "Tier Label", "Confidence",
            "Accruals", "Revenue Quality", "Capex/Life", "10-K Diff",
            "Call NLP", "Catalyst",
            "Top Red Flag 1", "Top Red Flag 2", "Top Red Flag 3",
            "Next Action",
        ]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        col_widths = [18, 8, 14, 6, 14, 10, 10, 14, 12, 10, 10, 10, 40, 40, 40, 50]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    TIER_COLORS = {1: "FF4444", 2: "FF9900", 3: "FFEB9C", 4: "C6EFCE"}
    now_str = _dt.now().strftime("%Y-%m-%d %H:%M")

    for res in results:
        sc = calculate_forensic_score(res)
        orch = res.reports.get("Orchestrator")
        orch_summary = orch.summary if orch and not orch.error else {}

        sub = sc["sub_scores"]
        red_flags = sc["red_flags"]

        def _flag(i: int) -> str:
            if i < len(red_flags):
                f = red_flags[i]
                return f"{f.get('agent','')}: {f.get('flag','')}"[:80]
            return ""

        next_actions = orch_summary.get("next_action", {}).get("actions", [])
        next_action_str = " | ".join(next_actions[:2]) if next_actions else ""

        row: list[Any] = [
            now_str, res.ticker,
            sc["forensic_score"], sc["tier"], sc["tier_label"],
            orch_summary.get("confidence", ""),
            sub.get("accruals_score", {}).get("score"),
            sub.get("revenue_quality_score", {}).get("score"),
            sub.get("capex_score", {}).get("score"),
            sub.get("tenk_diff_score", {}).get("score"),
            sub.get("call_nlp_score", {}).get("score"),
            sub.get("catalyst_score", {}).get("score"),
            _flag(0), _flag(1), _flag(2),
            next_action_str,
        ]
        ws.append(row)

        row_num = ws.max_row
        tier_val = sc["tier"]
        color = TIER_COLORS.get(tier_val, "FFFFFF")
        for col in (3, 4, 5):
            cell = ws.cell(row_num, col)
            cell.fill = PatternFill("solid", fgColor=color)
            if tier_val <= 2:
                cell.font = Font(bold=True, color="FFFFFF")

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# 전체 파이프라인: Phase 0 → Phase 0.5 → Phase 1 → Phase 2
# ---------------------------------------------------------------------------

async def analyze_stock(
    ticker: str,
    skip_orchestrator: bool = False,
    transcripts: list[str] | None = None,
) -> "ForensicResult":
    """6개 포렌식 에이전트 병렬 실행 → Opus 총괄.

    Session 4: Agent 1-3 Python-first 정량 메트릭
    Session 5: Agent 4 Language Diff Memo
    Session 6: Agent 5 Earnings Call NLP + Agent 6 Catalyst Monitor

    Phase 0:   Peer set 조회 + 기준 종목 & Peer XBRL + SEC 이벤트 데이터 병렬 fetch
    Phase 0.5: 모든 사전 계산 (Agent 1-6) 병렬 실행
    Phase 1:   6개 포렌식 에이전트 병렬 실행 (사전 계산 컨텍스트 포함)
    Phase 2:   Opus 총괄

    Args:
        ticker:            미국 종목 티커 (예: NVDA, MSFT, SMCI)
        skip_orchestrator: True면 Opus 총괄 단계 스킵 (빠른 실행)
        transcripts:       Earnings call transcript 텍스트 리스트 (없으면 SEC 8-K 사용)
    """
    started = time.perf_counter()
    ticker = ticker.upper().strip()

    # -----------------------------------------------------------------
    # Phase 0: Peer Set 조회 + 기준 종목 & Peer XBRL + SEC 이벤트 데이터 병렬 fetch
    # -----------------------------------------------------------------
    print(f"  🔍 [{ticker}] Phase 0: Peer set 조회 + 데이터 병렬 fetch …", flush=True)

    sector = get_sector_group(ticker)
    peers  = await get_peer_set(ticker, sector)
    print(f"    Sector: {sector} | Peers: {peers or '없음'}", flush=True)

    # XBRL + SEC 이벤트 데이터 동시 fetch
    fetch_tasks: dict[str, Any] = {
        "__self__":       fetch_single_xbrl(ticker, years=5),
        "__8k__":         ds.sec_8k_items(ticker, items=["4.02","4.01","5.02","8.01","2.06"], days=365),
        "__corresp__":    ds.sec_corresp(ticker, days=365),
        "__form4__":      ds.sec_form4(ticker, days=90),
        "__earnings__":   ds.sec_earnings_releases(ticker, quarters=6),
    }
    for p in peers:
        fetch_tasks[p] = fetch_single_xbrl(p, years=3)

    fetch_results = await asyncio.gather(
        *fetch_tasks.values(), return_exceptions=True
    )
    fetch_map = dict(zip(fetch_tasks.keys(), fetch_results))

    # XBRL 결과 분리
    self_ts = fetch_map.get("__self__")
    if isinstance(self_ts, Exception) or self_ts is None:
        print(f"    ⚠ [{ticker}] 기준 XBRL fetch 실패 — precomputed 메트릭 없이 진행", flush=True)
        self_ts = None

    peer_ts_map: dict[str, dict] = {
        p: v for p, v in fetch_map.items()
        if p not in ("__self__","__8k__","__corresp__","__form4__","__earnings__")
        and isinstance(v, dict)
    }

    # SEC 이벤트 데이터 분리 (Exception이면 None 처리)
    def _safe_get(key: str) -> Any:
        val = fetch_map.get(key)
        return None if isinstance(val, Exception) else val

    raw_8k      = _safe_get("__8k__")
    raw_corresp = _safe_get("__corresp__")
    raw_form4   = _safe_get("__form4__")
    raw_earnings = _safe_get("__earnings__")

    print(
        f"    ✓ 데이터 수집 완료: XBRL={'OK' if self_ts else 'FAIL'}, "
        f"peers={len(peer_ts_map)}/{len(peers)}, "
        f"8K={'OK' if raw_8k else '-'}, "
        f"CORRESP={'OK' if raw_corresp else '-'}, "
        f"Form4={'OK' if raw_form4 else '-'}",
        flush=True,
    )

    # -----------------------------------------------------------------
    # Phase 0.5: 6개 에이전트 전용 사전 계산 (Python, 병렬)
    # -----------------------------------------------------------------
    print(f"  ⚙  [{ticker}] Phase 0.5: 전체 사전 계산 …", flush=True)

    # --- Agent 1-3: XBRL 정량 메트릭 ---
    precomputed_map: dict[str, dict | None] = {}
    if self_ts:
        for key in ("accruals", "revenue", "capex"):
            try:
                precomputed_map[key] = compute_precomputed_context(key, self_ts, peer_ts_map)
                flags_count = _count_flags(precomputed_map[key])
                print(f"    ✓ {key}: {flags_count}개 flag", flush=True)
            except Exception as e:
                print(f"    ⚠ {key} 계산 실패: {e}", flush=True)
                precomputed_map[key] = None
    else:
        precomputed_map = {"accruals": None, "revenue": None, "capex": None}

    # --- Agent 4: Language Diff Memo (async, SEC fetch 포함) ---
    peer_context_for_diff = {"peers": peers, "sector": sector}
    diff_memo_task = compute_diff_memo(ticker, peer_context=peer_context_for_diff)

    # --- Agent 5: Earnings Call 사전 계산 ---
    def _compute_call_metrics() -> dict[str, Any]:
        try:
            result = agent5_precomputed(
                ticker=ticker,
                transcripts=transcripts or None,
                earnings_releases_raw=raw_earnings,
            )
            qtrs = result.get("quarters_analyzed", 0)
            flags_str = ", ".join(
                k for k, v in result.get("flags", {}).items()
                if v and v is not False and isinstance(v, bool)
            ) or "없음"
            print(f"    ✓ call_nlp: {qtrs}분기  flags=[{flags_str}]", flush=True)
            return result
        except Exception as e:
            print(f"    ⚠ call_nlp 계산 실패: {e}", flush=True)
            return {}

    # --- Agent 6: Catalyst Monitor 사전 계산 ---
    def _compute_catalyst_metrics() -> dict[str, Any]:
        try:
            result = agent6_precomputed(
                ticker=ticker,
                raw_8k=raw_8k,
                raw_corresp=raw_corresp,
                raw_form4=raw_form4,
                lookback_days=365,
            )
            prob  = result.get("catalyst_probability", "NONE")
            max_s = result.get("max_severity", 0)
            print(f"    ✓ catalyst: prob={prob}  max_severity={max_s}", flush=True)
            return result
        except Exception as e:
            print(f"    ⚠ catalyst 계산 실패: {e}", flush=True)
            return {}

    # Agent 4(async) + Agent 5/6(sync) 동시 실행
    diff_memo, call_result, catalyst_result = await asyncio.gather(
        diff_memo_task,
        asyncio.to_thread(_compute_call_metrics),
        asyncio.to_thread(_compute_catalyst_metrics),
    )

    call_precomputed     = call_result     if call_result     else None
    catalyst_precomputed = catalyst_result if catalyst_result else None

    # -----------------------------------------------------------------
    # Phase 1: 6개 에이전트 병렬 실행
    # -----------------------------------------------------------------
    agent_keys = list(AGENT_REGISTRY.keys())
    print(
        f"  ⚡ [{ticker}] Phase 1: {len(agent_keys)}개 포렌식 에이전트 병렬 실행 …",
        flush=True,
    )

    tasks = []
    for k in agent_keys:
        precomputed  = precomputed_map.get(k)
        memo         = diff_memo          if k == "tenk_diff" else None
        call_ctx     = call_precomputed   if k == "call_nlp"  else None
        catalyst_ctx = catalyst_precomputed if k == "catalyst" else None
        tasks.append(run_single_agent(
            k, ticker,
            precomputed=precomputed,
            diff_memo=memo,
            call_precomputed=call_ctx,
            catalyst_precomputed=catalyst_ctx,
        ))

    phase1_reports = await asyncio.gather(*tasks)
    reports: dict[str, AgentReport] = {r.agent: r for r in phase1_reports}

    for r in phase1_reports:
        status = "✓" if not r.error else "✗"
        score_val = _get_score_from_summary(r.agent, r.summary)
        print(
            f"    {status} {r.agent:<12} {r.elapsed_sec:.1f}s  score={score_val}",
            flush=True,
        )

    # -----------------------------------------------------------------
    # Phase 2: Opus 총괄
    # -----------------------------------------------------------------
    if not skip_orchestrator:
        interim = ForensicResult(
            ticker=ticker,
            elapsed_sec=time.perf_counter() - started,
            reports=reports,
        )
        print(f"  🔬 [{ticker}] Phase 2: Opus 총괄 분석 …", flush=True)
        orch_report = await run_orchestrator(interim)
        reports["Orchestrator"] = orch_report

        fs = calculate_forensic_score(ForensicResult(ticker, 0, reports))
        orch_summary = orch_report.summary
        final_score = orch_summary.get("forensic_score", fs["forensic_score"])
        final_tier  = orch_summary.get("tier_label", fs["tier_label"])
        print(
            f"    ✓ Orchestrator  {orch_report.elapsed_sec:.1f}s  "
            f"Score={final_score}  Tier={final_tier}",
            flush=True,
        )

    return ForensicResult(
        ticker=ticker,
        elapsed_sec=time.perf_counter() - started,
        reports=reports,
    )


async def analyze_stocks(
    tickers: list[str],
    skip_orchestrator: bool = False,
    save_excel: bool = False,
    excel_path: str | None = None,
    transcripts_map: dict[str, list[str]] | None = None,
) -> list["ForensicResult"]:
    """여러 종목 동시 분석 (종목 간 병렬).

    각 종목은 독립적으로 Phase 0~2 전체를 실행.

    Args:
        tickers:          분석 대상 티커 리스트
        skip_orchestrator: True면 Opus 총괄 단계 스킵
        save_excel:       True면 결과를 Excel 파일에 저장
        excel_path:       Excel 파일 경로 (없으면 기본 경로)
        transcripts_map:  {ticker: [transcript1, transcript2, ...]} 형식
                          없으면 SEC 8-K 자동 사용
    """
    print(
        f"🚀 {len(tickers)}개 종목 동시 포렌식 분석 시작: {', '.join(tickers)}",
        flush=True,
    )
    transcripts_map = transcripts_map or {}
    tasks = [
        analyze_stock(
            t,
            skip_orchestrator=skip_orchestrator,
            transcripts=transcripts_map.get(t),
        )
        for t in tickers
    ]
    results = list(await asyncio.gather(*tasks))

    if save_excel and not skip_orchestrator:
        path = export_to_excel(results, excel_path)
        print(f"\n💾 Excel 저장 완료: {path}", flush=True)

    return results


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------

def _count_flags(precomputed: dict | None) -> int:
    """사전 계산 딕셔너리에서 flag=True 항목 수 계산."""
    if not precomputed:
        return 0
    count = 0
    for v in precomputed.values():
        if isinstance(v, dict):
            if v.get("flag") is True:
                count += 1
            # 중첩 dict (latest 등)
            latest = v.get("latest")
            if isinstance(latest, dict) and latest.get("flag") is True:
                count += 1
    return count


def _get_score_from_summary(agent: str, summary: dict) -> Any:
    """에이전트별 score 키 자동 탐색."""
    key_order = [
        f"{agent}_score",
        "accruals_score", "revenue_quality_score", "capex_score",
        "tenk_diff_score", "call_nlp_score", "catalyst_score",
    ]
    for key in key_order:
        val = summary.get(key)
        if val is not None:
            return val
    return "?"
